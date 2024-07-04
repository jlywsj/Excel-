import pandas as pd
import os
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side


# 处理表格读取获取DataFrame后无效的数据
def handle_data(site: str, data_frame: pd.DataFrame) -> pd.DataFrame:
    # 添加新的列
    data_frame['Column7'] = pd.Series(dtype=str, index=data_frame.columns)
    data_frame['Column7'].fillna(site, inplace=True)
    # 将NaN值替换为空白，inplace表示在当前对象中操作（原地更改），不返回新的对象
    data_frame.fillna("", inplace=True)
    # 因为第一列数据是表头了，无效数据所以删除后返回
    return data_frame


# 用于将分解的多个列合并成一个新的DataFrame
def combine_dataframe(*args: pd.DataFrame) -> pd.DataFrame:
    # 创建一个空DataFrame等待接收合并
    combine_df = pd.DataFrame()
    # 迭代处理每个传入的 DataFrame 参数
    for arg in args:
        # 合并每个 DataFrame 到 combined_df，concat是pandas的函数，ignore_index表示是否忽略并建立新的索引（序号）
        combine_df = pd.concat([combine_df, arg], ignore_index=True)
    return combine_df


# 合并表格的主要函数
def merge_excels(file_paths, output_file):
    # 设置需要的列
    header = ['Column1', 'Column2', 'Column3', 'Column4', 'Column5', 'Column6', 'Column7']
    # 创建一个空的DataFrame对象用于接收读取的数据
    combined_df = pd.DataFrame(columns=header)
    # 遍历文件
    for file_path in file_paths:
        if os.path.isfile(file_path) and file_path.endswith('.xlsx'):
            # 从表格文件中读取DataFrame数据
            temp_df = pd.read_excel(file_path)

            # 处理表格文件中，不规范的部分
            temp_df = handle_data("Col7", temp_df)

            # 将需要的列合并为新的DataFrame
            temp_df = combine_dataframe(temp_df.loc[:, header])

            # 将数据合并到combined_df中
            combined_df = pd.concat([combined_df, temp_df], ignore_index=True)
        else:
            print(f"文件 {file_path} 不存在或不是一个Excel文件。")
    # 设置索引从1开始
    combined_df.index = combined_df.index + 1
    # 写入文件中去, index表示不写入索引, index_label设置索引列的列名
    combined_df.to_excel(output_file, index=True, sheet_name='Sheet1', index_label='序号')
    print(f"合并完成，结果保存在 '{output_file}' 文件中。")


# 获取文件列表
def get_files(path: str) -> list:
    files_path = []
    # os.walk(dir_path)：会递归地遍历指定文件夹下所有的层级目录
    # 每一次递归，返回root、 dirs、 files,分别表示当前的目录路径、当前目录路径包含的子目录、当前目录路径包含的文件
    for root, dirs, files in os.walk(path):
        # 遍历当前层级的文件
        for file in files:
            # 判断是否是要收集的文件
            if file.endswith(".xlsx"):
                files_path.append(os.path.join(root, file))
    return files_path


# 美化表格的样式
def prettify_page(file):
    # 加载现有Excel文件
    wb = load_workbook(file)

    # 获取默认的工作表
    ws = wb.active

    # 设置无框线样式
    thin_border = Border(left=Side(style='none'),
                         right=Side(style='none'),
                         top=Side(style='none'),
                         bottom=Side(style='none'))

    # 设置字体大小为12
    font = Font(size=12, bold=False)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.font = font
            cell.border = thin_border

    # 设置单元格对齐方式为居中
    alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = alignment

    # 设置行高为21
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 21

    # 设置列宽
    column_widths = {'A': 8, 'B': 10, 'C': 10, 'D': 10, 'E': 10, 'F': 10}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 保存修改后的工作簿
    wb.save(file)


# 从配置文件中读取文件路径，和输出的文件路径
def read_config():
    # 读取配置文件
    with open('config.yaml', 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
        return config['path']['read_path'], config['path']['output_path']


if __name__ == '__main__':
    # 读取文件配置
    READ_PATH, OUTPUT_FILE = read_config()
    FILE_PATHS = get_files(READ_PATH)
    merge_excels(FILE_PATHS, OUTPUT_FILE)
    # 修改样式
    prettify_page(OUTPUT_FILE)
